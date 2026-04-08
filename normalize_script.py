import os
import json
import pandas as pd
from openpyxl.styles import PatternFill, Font

# -----------------------------------
# CONFIGURATION
# -----------------------------------
INPUT_DIR = "data/model_csvs"
OUTPUT_DIR = "data/normalized_model_csvs"
STATS_FILE = "data/normalized_model_csvs/normalization_stats.json"

FEATURE_COLUMNS = [
    "FLOPs (G)",
    "Param Memory (MB)",
    "Activation Size (MB)",
    "pi_execution_time",
    "gpu_execution_time"
]

EPS = 1e-8
# -----------------------------------

os.makedirs(OUTPUT_DIR, exist_ok=True)

# -----------------------------------
# 1. Load all model CSVs
# -----------------------------------
dfs = []
file_names = []

for file in os.listdir(INPUT_DIR):
    if file.endswith(".csv"):
        path = os.path.join(INPUT_DIR, file)
        df = pd.read_csv(path)

        # Sanity check: required columns exist
        missing = [c for c in FEATURE_COLUMNS if c not in df.columns]
        if missing:
            raise ValueError(f"{file} missing columns: {missing}")

        dfs.append(df)
        file_names.append(file)

assert len(dfs) > 0, "❌ No CSV files found in data/model_csvs"

print(f"✔ Loaded {len(dfs)} model CSV files")

# -----------------------------------
# 2. Compute GLOBAL min / max
# -----------------------------------
concat_df = pd.concat(dfs, ignore_index=True)

global_min = {}
global_max = {}

for col in FEATURE_COLUMNS:
    global_min[col] = concat_df[col].min()
    global_max[col] = concat_df[col].max()

# -----------------------------------
# 3. Save normalization statistics
# -----------------------------------
stats = {
    col: {
        "min": float(global_min[col]),
        "max": float(global_max[col])
    }
    for col in FEATURE_COLUMNS
}

with open(STATS_FILE, "w") as f:
    json.dump(stats, f, indent=4)

print(f"✔ Saved normalization stats → {STATS_FILE}")

# -----------------------------------
# 4. Normalize each CSV and save
# -----------------------------------
column_colors = {
    "FLOPs (G)": "D9EAF7",            # light blue
    "Param Memory (MB)": "FCE4D6",    # light orange
    "Activation Size (MB)": "E2F0D9", # light green
    "pi_execution_time": "FFF2CC",    # light yellow
    "gpu_execution_time": "E4DFEC"    # light purple
}

for df, fname in zip(dfs, file_names):
    norm_df = df.copy()

    for col in FEATURE_COLUMNS:
        norm_df[col] = (
            (df[col] - global_min[col]) /
            (global_max[col] - global_min[col] + EPS)
        )

    # Save CSV
    csv_path = os.path.join(OUTPUT_DIR, fname)
    norm_df.to_csv(csv_path, index=False)

    # Save colored Excel
    excel_path = csv_path.replace(".csv", ".xlsx")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        norm_df.to_excel(writer, index=False, sheet_name="Normalized")
        worksheet = writer.sheets["Normalized"]

        # Apply column colors
        for col_idx, col_name in enumerate(norm_df.columns, 1):
            if col_name in column_colors:
                fill = PatternFill(
                    fill_type="solid",
                    fgColor=column_colors[col_name]
                )

                for row in range(1, len(norm_df) + 2):
                    worksheet.cell(row=row, column=col_idx).fill = fill

        # Make headers bold
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    print(f"✔ Saved normalized CSV → {csv_path}")
    print(f"✔ Saved colored Excel → {excel_path}")

print("\n✅ All model CSVs normalized successfully.")